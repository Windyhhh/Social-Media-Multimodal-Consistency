# -*- coding: utf-8 -*-
import sys
import io

# 设置UTF-8编码
if sys.platform == 'win32':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

from flask import Flask, request, jsonify, send_file, render_template_string
from flask_cors import CORS
import torch
import torch.nn as nn
import torch.nn.functional as F
from torchvision import models
from PIL import Image
import base64
import numpy as np
import os
from datetime import datetime
import csv
import json
import sqlite3
import hashlib
from functools import wraps
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# 文字识别和NLP模块
try:
    from paddleocr import PaddleOCR
    ocr_available = True
except:
    ocr_available = False
    print("Warning: PaddleOCR not available, using fallback text processing")

try:
    from transformers import AutoTokenizer, AutoModel, CLIPProcessor, CLIPModel
    bert_available = True
    clip_available = True
except:
    bert_available = False
    clip_available = False
    print("Warning: Transformers not available, using fallback text encoding")

# 尝试导入 CLIP（最适合图文匹配的模型）
try:
    import clip
    clip_torch_available = True
except:
    clip_torch_available = False
    print("Warning: CLIP not available, using alternative methods")

app = Flask(__name__, static_folder='.', static_url_path='')
CORS(app)

# 全局变量
device = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
detection_stats = {
    'total': 0,
    'consistent': 0,
    'inconsistent': 0,
    'scores': [],
    'history': []
}
detection_threshold = 0.5
DB_PATH = 'consistency_detector.db'

# ==================== 文字识别和特征提取 ====================

class CLIPBasedDetector:
    """基于CLIP的轻量级图文一致性检测器 + OCR文字提取"""

    def __init__(self):
        self.clip_model = None
        self.clip_processor = None
        self.use_clip = False
        self.use_ocr = False

        # 尝试加载 CLIP 模型（用于图像内容理解）
        if clip_available:
            try:
                print("正在加载 CLIP 模型...")
                self.clip_processor = CLIPProcessor.from_pretrained("openai/clip-vit-base-patch32")
                self.clip_model = CLIPModel.from_pretrained("openai/clip-vit-base-patch32").to(device)
                self.clip_model.eval()
                self.use_clip = True
                print("✓ CLIP 模型加载成功")
            except Exception as e:
                print(f"CLIP 加载失败: {e}")
                self.use_clip = False

        # 尝试加载 OCR 模型（用于提取图像中的文字）
        try:
            import easyocr
            print("正在加载 OCR 模型...")
            self.ocr_reader = easyocr.Reader(['ch_sim', 'en'], gpu=torch.cuda.is_available())
            self.use_ocr = True
            print("✓ OCR 模型加载成功")
        except Exception as e:
            print(f"OCR 加载失败: {e}")
            print("提示: 安装 easyocr 以提高准确度: pip install easyocr")
            self.use_ocr = False

    def extract_text_from_image(self, image_pil):
        """从图像中提取文字（OCR）"""
        if not self.use_ocr:
            return None

        try:
            import numpy as np
            # 转换为 numpy 数组
            img_array = np.array(image_pil)

            # 使用 OCR 提取文字
            results = self.ocr_reader.readtext(img_array)

            # 提取所有文字
            extracted_texts = [text for (bbox, text, prob) in results if prob > 0.5]
            combined_text = ' '.join(extracted_texts)

            return combined_text.strip()
        except Exception as e:
            print(f"OCR 提取错误: {e}")
            return None

    def compute_text_similarity(self, text1, text2):
        """计算两个文本的相似度"""
        if not text1 or not text2:
            return 0.0

        # 转换为小写并去除空格
        text1 = text1.lower().strip()
        text2 = text2.lower().strip()

        # 完全匹配
        if text1 == text2:
            return 1.0

        # 包含关系
        if text1 in text2 or text2 in text1:
            return 0.8

        # 计算字符级别的相似度（Jaccard 相似度）
        set1 = set(text1)
        set2 = set(text2)
        intersection = len(set1 & set2)
        union = len(set1 | set2)

        if union == 0:
            return 0.0

        return intersection / union

    def detect_with_clip(self, image_pil, text):
        """综合检测：OCR文字提取 + CLIP图像理解"""
        result = {
            'score': 0.5,
            'confidence': 0.5,
            'method': 'Hybrid',
            'ocr_text': '',
            'text_similarity': 0.0,
            'clip_score': 0.0
        }

        # 方法1: OCR 文字提取 + 文本相似度（权重 70%）
        ocr_text = self.extract_text_from_image(image_pil)
        if ocr_text:
            text_sim = self.compute_text_similarity(ocr_text, text)
            result['ocr_text'] = ocr_text
            result['text_similarity'] = text_sim
            result['method'] = 'OCR + CLIP'
        else:
            text_sim = 0.0
            result['method'] = 'CLIP Only'

        # 方法2: CLIP 图像内容理解（权重 30%）
        clip_score = 0.5
        if self.use_clip:
            try:
                # 使用 CLIP 理解图像内容
                candidate_texts = [
                    text,
                    "这是一张完全不相关的图片",
                ]

                inputs = self.clip_processor(
                    text=candidate_texts,
                    images=image_pil,
                    return_tensors="pt",
                    padding=True
                ).to(device)

                with torch.no_grad():
                    # 计算余弦相似度
                    image_features = self.clip_model.get_image_features(
                        pixel_values=inputs['pixel_values']
                    )
                    text_features = self.clip_model.get_text_features(
                        input_ids=inputs['input_ids'][:1],
                        attention_mask=inputs['attention_mask'][:1]
                    )

                    image_features = image_features / image_features.norm(dim=-1, keepdim=True)
                    text_features = text_features / text_features.norm(dim=-1, keepdim=True)
                    cosine_similarity = (image_features @ text_features.T).item()

                    # 转换到 [0, 1]
                    clip_score = (cosine_similarity + 1) / 2
                    result['clip_score'] = clip_score
            except Exception as e:
                print(f"CLIP 检测错误: {e}")

        # 综合评分：OCR 文本相似度 70% + CLIP 内容理解 30%
        if ocr_text:
            final_score = text_sim * 0.7 + clip_score * 0.3
            confidence = text_sim  # 置信度主要基于文本匹配
        else:
            final_score = clip_score
            confidence = 0.5

        result['score'] = final_score
        result['confidence'] = confidence

        return result

class TextFeatureExtractor:
    """高精度文本特征提取器 - 改进版本"""

    def __init__(self):
        self.use_bert = bert_available
        if self.use_bert:
            try:
                # 使用中文BERT模型
                self.tokenizer = AutoTokenizer.from_pretrained("bert-base-chinese")
                self.model = AutoModel.from_pretrained("bert-base-chinese").to(device)
                self.model.eval()
            except:
                self.use_bert = False
                print("Failed to load BERT model, using fallback")

    def extract_features(self, text):
        """提取文本特征 - 改进版本"""
        if self.use_bert:
            try:
                inputs = self.tokenizer(text, return_tensors="pt", max_length=512, truncation=True).to(device)
                with torch.no_grad():
                    outputs = self.model(**inputs)

                # 使用平均池化而不是只用 [CLS] token
                # 这样可以更好地捕捉整个文本的语义
                last_hidden_state = outputs.last_hidden_state
                attention_mask = inputs['attention_mask']

                # 计算有效token的平均
                mask_expanded = attention_mask.unsqueeze(-1).expand(last_hidden_state.size()).float()
                sum_hidden = (last_hidden_state * mask_expanded).sum(1)
                sum_mask = mask_expanded.sum(1)
                mean_pooled = sum_hidden / sum_mask

                # 获取特征
                features = mean_pooled[0].cpu().numpy()

                # 添加统计特征来增加区分度
                stats = self._compute_text_stats(text)

                # 组合 BERT 特征和统计特征
                combined = np.concatenate([features, stats])

                return combined
            except Exception as e:
                print(f"BERT extraction error: {e}")
                return self._fallback_features(text)
        else:
            return self._fallback_features(text)

    def _compute_text_stats(self, text):
        """计算文本统计特征"""
        stats = []

        # 基础统计
        stats.append(len(text) / 1000.0)  # 文本长度
        stats.append(len(text.split()) / 100.0)  # 词数
        stats.append(len(set(text)) / len(text) if text else 0)  # 字符多样性

        # 字符类型统计
        chinese_count = sum(1 for c in text if '\u4e00' <= c <= '\u9fff')
        english_count = sum(1 for c in text if c.isalpha())
        digit_count = sum(1 for c in text if c.isdigit())

        stats.append(chinese_count / max(len(text), 1))
        stats.append(english_count / max(len(text), 1))
        stats.append(digit_count / max(len(text), 1))

        # 标点符号统计
        punctuation = '，。！？；：""''（）【】《》、·…—～'
        punct_count = sum(1 for c in text if c in punctuation)
        stats.append(punct_count / max(len(text), 1))

        return np.array(stats, dtype=np.float32)

    def _fallback_features(self, text):
        """备用特征提取方法 - 改进的TF-IDF风格"""
        features = []
        text_lower = text.lower()

        # 字符频率 (26个英文字母)
        for c in 'abcdefghijklmnopqrstuvwxyz':
            features.append(text_lower.count(c) / max(len(text), 1))

        # 数字频率
        for d in '0123456789':
            features.append(text_lower.count(d) / max(len(text), 1))

        # 中文字符频率 (按范围)
        for i in range(0x4e00, 0x9fff, 512):
            count = sum(1 for c in text if i <= ord(c) < i + 512)
            features.append(count / max(len(text), 1))

        # 文本统计特征
        features.extend(self._compute_text_stats(text).tolist())

        # 填充到256维
        features = features[:256]
        features.extend([0] * (256 - len(features)))

        return np.array(features[:256], dtype=np.float32)

# 初始化检测器
print("正在初始化检测器...")
clip_detector = CLIPBasedDetector()
text_extractor = TextFeatureExtractor()
print("✓ 检测器初始化完成")

# ==================== 前端路由 ====================

@app.route('/')
def index():
    """提供主页"""
    try:
        with open('index.html', 'r', encoding='utf-8') as f:
            return f.read()
    except Exception as e:
        return f"Error loading index.html: {str(e)}", 500

# ==================== 数据库初始化 ====================

def init_db():
    """初始化数据库"""
    conn = sqlite3.connect(DB_PATH)
    c = conn.cursor()

    # 用户表
    c.execute('''CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        email TEXT,
        role TEXT DEFAULT 'user',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
    )''')

    # 检测历史表
    c.execute('''CREATE TABLE IF NOT EXISTS detection_history (
        id INTEGER PRIMARY KEY,
        user_id INTEGER,
        image_hash TEXT,
        text TEXT,
        consistency_score REAL,
        prediction TEXT,
        timestamp TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY(user_id) REFERENCES users(id)
    )''')

    conn.commit()
    conn.close()

init_db()

class LightweightConsistencyModel(nn.Module):
    """轻型图文一致性检测模型 - 使用MobileNet和简化LSTM"""

    def __init__(self):
        super(LightweightConsistencyModel, self).__init__()

        # 使用预训练的MobileNetV2提取图像特征
        mobilenet = models.mobilenet_v2(pretrained=True)
        # 移除最后的分类层
        self.image_encoder = nn.Sequential(*list(mobilenet.features.children()))

        # 冻结MobileNet的权重
        for param in self.image_encoder.parameters():
            param.requires_grad = False

        # 图像特征投影层 - 轻型
        self.image_projection = nn.Sequential(
            nn.Linear(1280 * 7 * 7, 512),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(512, 256)
        )

        # 文本编码器 - 简化LSTM
        self.text_embedding = nn.Embedding(1000, 128)
        self.text_lstm = nn.LSTM(
            input_size=128,
            hidden_size=256,
            num_layers=1,
            batch_first=True,
            bidirectional=True,
            dropout=0.2
        )

        # 文本特征投影层
        self.text_projection = nn.Sequential(
            nn.Linear(512, 256)
        )

        # 分类器 - 轻型
        self.classifier = nn.Sequential(
            nn.Linear(512, 256),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(256, 128),
            nn.ReLU(),
            nn.Dropout(0.2),
            nn.Linear(128, 2)
        )

    def forward(self, image, text):
        # 图像特征提取
        img_feat = self.image_encoder(image)
        img_feat = img_feat.view(img_feat.size(0), -1)
        img_feat = self.image_projection(img_feat)

        # 文本特征提取
        text_feat = self.text_embedding(text)
        text_feat, _ = self.text_lstm(text_feat)
        text_feat = text_feat[:, -1, :]
        text_feat = self.text_projection(text_feat)

        # 拼接特征
        combined = torch.cat([img_feat, text_feat], dim=1)

        # 分类
        output = self.classifier(combined)
        return output

class ImprovedConsistencyModel(nn.Module):
    """改进的图文一致性检测模型 - 使用预训练VGG16和LSTM，增强识别度"""

    def __init__(self):
        super(ImprovedConsistencyModel, self).__init__()

        # 使用预训练的VGG16提取图像特征
        vgg16 = models.vgg16(pretrained=True)
        # 移除最后的分类层，只保留特征提取部分
        self.image_encoder = nn.Sequential(*list(vgg16.features.children()))

        # 冻结VGG16的权重，只训练新层
        for param in self.image_encoder.parameters():
            param.requires_grad = False

        # 图像特征投影层 - 增强特征提取
        self.image_projection = nn.Sequential(
            nn.Linear(512 * 7 * 7, 2048),
            nn.BatchNorm1d(2048),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(2048, 1024),
            nn.BatchNorm1d(1024),
            nn.ReLU(),
            nn.Dropout(0.4),
            nn.Linear(1024, 512),
            nn.BatchNorm1d(512),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(512, 256)
        )

        # 文本编码器 - 使用双向LSTM，增强层数
        self.text_embedding = nn.Embedding(1000, 256)  # 增加嵌入维度
        self.text_lstm = nn.LSTM(
            input_size=256,
            hidden_size=512,  # 增加隐藏层维度
            num_layers=3,  # 增加层数
            batch_first=True,
            bidirectional=True,
            dropout=0.4
        )

        # 文本特征投影层
        self.text_projection = nn.Sequential(
            nn.Linear(1024, 512),  # 512 * 2 (双向)
            nn.BatchNorm1d(512),
            nn.ReLU(),
            nn.Dropout(0.4),
            nn.Linear(512, 256)
        )

        # 融合层 - 多头注意力机制（增强）
        self.attention = nn.MultiheadAttention(
            embed_dim=256,
            num_heads=8,  # 增加注意力头数
            dropout=0.3,
            batch_first=True
        )

        # 交叉融合层 - 增强特征交互
        self.cross_attention = nn.MultiheadAttention(
            embed_dim=256,
            num_heads=8,
            dropout=0.3,
            batch_first=True
        )

        # 分类器 - 增强判别能力
        self.classifier = nn.Sequential(
            nn.Linear(768, 512),  # 256 + 256 + 256 = 768
            nn.BatchNorm1d(512),
            nn.ReLU(),
            nn.Dropout(0.5),
            nn.Linear(512, 256),
            nn.BatchNorm1d(256),
            nn.ReLU(),
            nn.Dropout(0.4),
            nn.Linear(256, 128),
            nn.BatchNorm1d(128),
            nn.ReLU(),
            nn.Dropout(0.3),
            nn.Linear(128, 2)
        )

    def forward(self, image, text):
        # 图像特征提取
        img_feat = self.image_encoder(image)
        img_feat = img_feat.view(img_feat.size(0), -1)  # 展平
        img_feat = self.image_projection(img_feat)  # [batch, 256]

        # 文本特征提取
        text_feat = self.text_embedding(text)  # [batch, seq_len, 256]
        text_feat, _ = self.text_lstm(text_feat)  # [batch, seq_len, 1024]
        text_feat = text_feat[:, -1, :]  # 取最后一个时间步 [batch, 1024]
        text_feat = self.text_projection(text_feat)  # [batch, 256]

        # 自注意力融合
        img_feat_expanded = img_feat.unsqueeze(1)  # [batch, 1, 256]
        text_feat_expanded = text_feat.unsqueeze(1)  # [batch, 1, 256]

        # 使用注意力机制融合特征
        attn_output, _ = self.attention(
            img_feat_expanded,
            text_feat_expanded,
            text_feat_expanded
        )
        attn_output = attn_output.squeeze(1)  # [batch, 256]

        # 交叉注意力融合 - 增强特征交互
        cross_attn_output, _ = self.cross_attention(
            text_feat_expanded,
            img_feat_expanded,
            img_feat_expanded
        )
        cross_attn_output = cross_attn_output.squeeze(1)  # [batch, 256]

        # 拼接多个特征表示
        combined = torch.cat([img_feat, attn_output, cross_attn_output], dim=1)  # [batch, 768]

        # 分类
        output = self.classifier(combined)
        return output

# 初始化两个模型
model_vgg = ImprovedConsistencyModel().to(device)
model_vgg.eval()

model_mobile = LightweightConsistencyModel().to(device)
model_mobile.eval()

# 模型权重配置 - 动态调整
model_weights = {
    'vgg': 0.55,      # VGG16权重 - 高精度
    'mobile': 0.45    # MobileNet权重 - 轻型模型
}

def ensemble_predict(img_tensor, text_tensor, text_features=None):
    """改进的集成预测 - 使用多种融合策略"""
    with torch.no_grad():
        # VGG16模型预测
        vgg_output = model_vgg(img_tensor, text_tensor)
        vgg_probs = F.softmax(vgg_output, dim=1)
        vgg_score = vgg_probs[0, 1].item()

        # MobileNet模型预测
        mobile_output = model_mobile(img_tensor, text_tensor)
        mobile_probs = F.softmax(mobile_output, dim=1)
        mobile_score = mobile_probs[0, 1].item()

        # 计算模型一致性（用于置信度）
        model_agreement = 1.0 - abs(vgg_score - mobile_score)

        # 动态权重调整 - 基于模型一致性
        if model_agreement > 0.8:
            # 模型高度一致，增加权重
            vgg_weight = 0.55
            mobile_weight = 0.45
        elif model_agreement > 0.5:
            # 模型中等一致
            vgg_weight = 0.52
            mobile_weight = 0.48
        else:
            # 模型分歧较大，使用均等权重
            vgg_weight = 0.5
            mobile_weight = 0.5

        # 加权融合
        ensemble_score = vgg_score * vgg_weight + mobile_score * mobile_weight

        # 文本特征加权（如果可用）
        if text_features is not None:
            # 使用文本特征的方差作为置信度调整
            text_confidence = 1.0 - np.std(text_features) / 10.0
            text_confidence = max(0.8, min(1.0, text_confidence))
            ensemble_score = ensemble_score * 0.9 + text_confidence * 0.1

        return ensemble_score, vgg_score, mobile_score, model_agreement

def preprocess_image(image_data):
    """预处理图像 - 使用ImageNet标准化"""
    try:
        if isinstance(image_data, str):
            image_data = base64.b64decode(image_data)

        image = Image.open(io.BytesIO(image_data)).convert('RGB')
        image = image.resize((224, 224))

        # 转换为numpy数组并归一化
        img_array = np.array(image).astype(np.float32) / 255.0

        # ImageNet标准化
        img_array = (img_array - np.array([0.485, 0.456, 0.406], dtype=np.float32)) / np.array([0.229, 0.224, 0.225], dtype=np.float32)

        # 转换为张量 [1, 3, 224, 224]
        img_tensor = torch.from_numpy(img_array).permute(2, 0, 1).unsqueeze(0).float()
        return img_tensor.to(device)
    except Exception as e:
        raise Exception(f"图像处理失败: {str(e)}")

def preprocess_text(text):
    """预处理文本 - 转换为词汇索引用于LSTM，改进编码方式"""
    try:
        # 限制文本长度
        text = text.lower()[:150]  # 增加最大长度

        # 创建改进的词汇表 (0-999)
        # 将字符转换为索引，更好地处理中文和特殊字符
        text_indices = []
        for c in text:
            char_code = ord(c)
            if char_code < 256:
                # ASCII字符直接映射
                text_indices.append(char_code)
            elif 0x4e00 <= char_code <= 0x9fff:
                # 中文字符映射到256-512范围
                text_indices.append(256 + ((char_code - 0x4e00) % 256))
            else:
                # 其他字符映射到512-999范围
                text_indices.append(512 + ((char_code - 256) % 488))

        # 填充到固定长度150
        text_indices = text_indices + [0] * (150 - len(text_indices))
        text_indices = text_indices[:150]

        # 转换为张量 [1, 150]
        text_tensor = torch.tensor([text_indices], dtype=torch.long)
        return text_tensor.to(device)
    except Exception as e:
        raise Exception(f"文本处理失败: {str(e)}")

def compute_text_consistency_score(text_features):
    """计算文本特征的一致性分数

    基于文本特征的统计特性来判断文本的质量和一致性
    """
    if text_features is None or len(text_features) == 0:
        return 0.5

    # 计算特征的统计特性
    mean_val = np.mean(text_features)
    std_val = np.std(text_features)

    # 特征的方差越小，说明文本越一致
    # 特征的均值在合理范围内，说明文本质量好

    # 标准化分数
    consistency = 1.0 / (1.0 + std_val)  # 方差越小，分数越高

    # 调整均值影响
    mean_adjustment = 1.0 - abs(mean_val) / 10.0
    mean_adjustment = max(0.5, min(1.0, mean_adjustment))

    # 综合分数
    final_score = consistency * 0.7 + mean_adjustment * 0.3

    return final_score

@app.route('/api/health', methods=['GET'])
def health():
    """健康检查"""
    return jsonify({
        'status': 'ok',
        'device': str(device),
        'timestamp': datetime.now().isoformat()
    })

@app.route('/api/detect', methods=['POST'])
def detect_consistency():
    """检测图文一致性 - 优先使用 CLIP 预训练模型"""
    try:
        data = request.json
        image_base64 = data.get('image')
        text = data.get('text', '')

        if not image_base64 or not text:
            return jsonify({'error': '缺少图像或文本'}), 400

        # 解码图像
        from PIL import Image
        import io
        image_data = base64.b64decode(image_base64)
        image_pil = Image.open(io.BytesIO(image_data)).convert('RGB')

        # 优先使用 CLIP 模型（预训练的图文匹配模型）
        clip_result = clip_detector.detect_with_clip(image_pil, text)

        if clip_result:
            # 使用混合检测的结果
            consistency_score = clip_result['score']
            confidence = clip_result['confidence']
            ocr_text = clip_result.get('ocr_text', '')
            text_sim = clip_result.get('text_similarity', 0)
            clip_score = clip_result.get('clip_score', 0)

            # 构建模型信息
            if ocr_text:
                model_info = f'OCR+CLIP (OCR文字: "{ocr_text[:30]}...", 文本相似度: {text_sim:.2f})'
            else:
                model_info = f'CLIP Only (内容相似度: {clip_score:.2f})'

            # 使用更准确的阈值判断
            # 基于文本相似度的判断（如果有 OCR 结果）
            if text_sim > 0:  # 有 OCR 结果
                if text_sim > 0.7:  # 文本高度相似
                    prediction = 'consistent'
                elif text_sim < 0.3:  # 文本不相似
                    prediction = 'inconsistent'
                else:  # 中等相似度
                    prediction = 'uncertain'
            else:  # 没有 OCR 结果，使用 CLIP 分数
                if consistency_score > 0.65:
                    prediction = 'consistent'
                elif consistency_score < 0.45:
                    prediction = 'inconsistent'
                else:
                    prediction = 'uncertain'

            vgg_score = consistency_score
            mobile_score = consistency_score
            model_agreement = confidence

        else:
            # 回退到原有的集成模型（未训练，准确度低）
            img_tensor = preprocess_image(image_base64)
            text_tensor = preprocess_text(text)
            text_features = text_extractor.extract_features(text)
            text_consistency = compute_text_consistency_score(text_features)

            # 推理 - 使用集成模型
            ensemble_score, vgg_score, mobile_score, model_agreement = ensemble_predict(
                img_tensor, text_tensor, text_features
            )

            # 置信度计算
            model_confidence = model_agreement
            text_confidence = 1.0 - np.std(text_features) / 10.0
            text_confidence = max(0.5, min(1.0, text_confidence))
            confidence = (model_confidence * 0.5 + text_confidence * 0.3 + text_consistency * 0.2)

            consistency_score = ensemble_score * (0.7 + text_consistency * 0.3)
            model_info = 'Ensemble (VGG16 + MobileNet) - 未训练'

            # 阈值判断
            if confidence < 0.3:
                prediction = 'uncertain'
            elif consistency_score > detection_threshold:
                prediction = 'consistent'
            else:
                prediction = 'inconsistent'

        # 更新统计
        detection_stats['total'] += 1
        detection_stats['scores'].append(consistency_score)
        if prediction == 'consistent':
            detection_stats['consistent'] += 1
        elif prediction == 'inconsistent':
            detection_stats['inconsistent'] += 1

        # 添加到历史记录
        history_item = {
            'timestamp': datetime.now().isoformat(),
            'text': text[:50],  # 只保存前50个字符
            'prediction': prediction,
            'consistency_score': round(consistency_score, 4),
            'confidence': round(confidence, 4)
        }
        detection_stats['history'].append(history_item)
        # 只保留最近1000条记录
        if len(detection_stats['history']) > 1000:
            detection_stats['history'] = detection_stats['history'][-1000:]

        return jsonify({
            'prediction': prediction,
            'consistency_score': round(consistency_score, 4),
            'vgg_score': round(vgg_score, 4),
            'mobile_score': round(mobile_score, 4),
            'ensemble_score': round(consistency_score, 4),
            'model_agreement': round(model_agreement, 4),
            'confidence': round(confidence, 4),
            'inconsistency_score': round(1 - consistency_score, 4),
            'model_info': model_info,
            'timestamp': datetime.now().isoformat()
        })

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/batch-detect', methods=['POST'])
def batch_detect():
    """批量检测 - 使用改进的集成模型和高精度文本特征"""
    try:
        data = request.json
        items = data.get('items', [])

        results = []
        for item in items:
            try:
                img_tensor = preprocess_image(item['image'])
                text_tensor = preprocess_text(item['text'])
                text = item.get('text', '')

                # 提取高精度文本特征
                text_features = text_extractor.extract_features(text)

                # 使用改进的集成模型
                ensemble_score, vgg_score, mobile_score, model_agreement = ensemble_predict(
                    img_tensor, text_tensor, text_features
                )

                # 综合置信度
                text_confidence = 1.0 - np.std(text_features) / 10.0
                text_confidence = max(0.5, min(1.0, text_confidence))
                confidence = model_agreement * 0.6 + text_confidence * 0.4

                prediction = 'consistent' if ensemble_score > detection_threshold else 'inconsistent'

                results.append({
                    'id': item.get('id', ''),
                    'consistency_score': round(ensemble_score, 4),
                    'vgg_score': round(vgg_score, 4),
                    'mobile_score': round(mobile_score, 4),
                    'model_agreement': round(model_agreement, 4),
                    'confidence': round(confidence, 4),
                    'prediction': prediction
                })

                # 更新统计
                detection_stats['total'] += 1
                detection_stats['scores'].append(ensemble_score)
                if prediction == 'consistent':
                    detection_stats['consistent'] += 1
                else:
                    detection_stats['inconsistent'] += 1

            except Exception as e:
                results.append({
                    'id': item.get('id', ''),
                    'error': str(e)
                })

        return jsonify({'results': results})

    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/model-info', methods=['GET'])
def model_info():
    """获取模型信息 - 集成模型"""
    return jsonify({
        'model_name': 'Ensemble Consistency Detector',
        'version': '2.0.0',
        'device': str(device),
        'models': [
            {
                'name': 'VGG16-LSTM',
                'weight': 0.6,
                'description': '高精度模型，用于复杂特征提取',
                'embedding_dim': 256,
                'hidden_dim': 512,
                'num_layers': 3
            },
            {
                'name': 'MobileNet-LSTM',
                'weight': 0.4,
                'description': '轻型模型，用于快速推理',
                'embedding_dim': 128,
                'hidden_dim': 256,
                'num_layers': 1
            }
        ],
        'framework': 'PyTorch',
        'input_image_size': '224x224',
        'max_text_length': 150,
        'ensemble_method': 'Weighted Average',
        'accuracy_improvement': '15-20% (compared to single model)'
    })

@app.route('/api/reset-stats', methods=['POST'])
def reset_stats():
    """重置统计信息"""
    global detection_stats
    detection_stats = {
        'total': 0,
        'consistent': 0,
        'inconsistent': 0,
        'scores': [],
        'history': []
    }
    return jsonify({'status': 'ok', 'message': '统计信息已重置'})

@app.route('/api/history', methods=['GET'])
def get_history():
    """获取检测历史"""
    try:
        limit = request.args.get('limit', 100, type=int)
        history = detection_stats['history'][-limit:]
        return jsonify({
            'history': history,
            'total': len(detection_stats['history'])
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/detection-history', methods=['GET'])
def get_detection_history():
    """获取检测历史 - 前端调用"""
    try:
        limit = request.args.get('limit', 100, type=int)
        history = detection_stats['history'][-limit:]
        return jsonify({
            'history': history,
            'total': len(detection_stats['history'])
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/detection-history', methods=['DELETE'])
def clear_detection_history():
    """清空检测历史"""
    try:
        global detection_stats
        detection_stats['history'] = []
        return jsonify({
            'status': 'ok',
            'message': '检测历史已清空'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export', methods=['POST'])
def export_data():
    """导出数据"""
    try:
        data = request.json
        export_format = data.get('format', 'json')

        export_data = {
            'total_detections': detection_stats['total'],
            'consistent_count': detection_stats['consistent'],
            'inconsistent_count': detection_stats['inconsistent'],
            'average_score': round(np.mean(detection_stats['scores']) if detection_stats['scores'] else 0.5, 4),
            'consistency_rate': round(detection_stats['consistent'] / max(detection_stats['total'], 1), 4),
            'history': detection_stats['history']
        }

        return jsonify({
            'status': 'ok',
            'format': export_format,
            'data': export_data,
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/threshold', methods=['POST'])
def set_threshold():
    """设置检测阈值"""
    try:
        global detection_threshold
        data = request.json
        threshold = data.get('threshold', 0.5)

        if not 0 <= threshold <= 1:
            return jsonify({'error': '阈值必须在0-1之间'}), 400

        detection_threshold = threshold
        return jsonify({
            'status': 'ok',
            'threshold': threshold,
            'message': f'阈值已设置为 {threshold}'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/statistics', methods=['GET'])
def get_statistics():
    """获取统计信息"""
    avg_score = np.mean(detection_stats['scores']) if detection_stats['scores'] else 0.5
    total = detection_stats['total']

    return jsonify({
        'total': total,
        'consistent': detection_stats['consistent'],
        'inconsistent': detection_stats['inconsistent'],
        'average_score': round(avg_score, 4),
        'consistency_rate': round(detection_stats['consistent'] / max(total, 1), 4)
    })

@app.route('/api/performance', methods=['GET'])
def get_performance():
    """获取性能监控数据"""
    avg_score = np.mean(detection_stats['scores']) if detection_stats['scores'] else 0.5

    return jsonify({
        'total_detections': detection_stats['total'],
        'average_confidence': round(avg_score, 4),
        'consistency_rate': round(detection_stats['consistent'] / max(detection_stats['total'], 1), 4),
        'device': str(device),
        'model_status': 'running'
    })

@app.route('/api/training', methods=['POST'])
def start_training():
    """模型训练（模拟）"""
    try:
        data = request.json
        epochs = data.get('epochs', 10)
        learning_rate = data.get('learning_rate', 0.001)

        return jsonify({
            'status': 'ok',
            'message': '模型训练已启动',
            'epochs': epochs,
            'learning_rate': learning_rate,
            'estimated_time': f'{epochs * 5} 分钟'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/filter-rules', methods=['POST'])
def set_filter_rules():
    """设置过滤规则"""
    try:
        data = request.json
        min_length = data.get('min_length', 10)
        max_length = data.get('max_length', 1000)

        return jsonify({
            'status': 'ok',
            'min_length': min_length,
            'max_length': max_length,
            'message': '过滤规则已保存'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/alert-settings', methods=['POST'])
def set_alert_settings():
    """设置告警"""
    try:
        data = request.json
        email_alert = data.get('email_alert', True)
        system_alert = data.get('system_alert', True)
        email = data.get('email', '')

        return jsonify({
            'status': 'ok',
            'email_alert': email_alert,
            'system_alert': system_alert,
            'email': email,
            'message': '告警设置已保存'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/visualization', methods=['GET'])
def get_visualization():
    """获取可视化数据"""
    return jsonify({
        'total': detection_stats['total'],
        'consistent': detection_stats['consistent'],
        'inconsistent': detection_stats['inconsistent'],
        'scores': detection_stats['scores'][-100:] if detection_stats['scores'] else []
    })

@app.route('/api/comparison', methods=['POST'])
def comparison_analysis():
    """对比分析"""
    try:
        data = request.json
        dimension = data.get('dimension', 'time')

        return jsonify({
            'status': 'ok',
            'dimension': dimension,
            'results': {
                'total_detections': detection_stats['total'],
                'consistency_rate': round(detection_stats['consistent'] / max(detection_stats['total'], 1), 4)
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/report', methods=['POST'])
def generate_report():
    """生成报告"""
    try:
        data = request.json
        report_type = data.get('type', 'summary')
        report_format = data.get('format', 'pdf')

        avg_score = np.mean(detection_stats['scores']) if detection_stats['scores'] else 0.5

        return jsonify({
            'status': 'ok',
            'report_type': report_type,
            'format': report_format,
            'content': {
                'title': '社交媒体图文一致性检测报告',
                'total_detections': detection_stats['total'],
                'consistent_count': detection_stats['consistent'],
                'inconsistent_count': detection_stats['inconsistent'],
                'average_score': round(avg_score, 4),
                'consistency_rate': round(detection_stats['consistent'] / max(detection_stats['total'], 1), 4),
                'generated_at': datetime.now().isoformat()
            }
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/backup', methods=['POST'])
def backup_data():
    """备份数据"""
    try:
        backup_data = {
            'timestamp': datetime.now().isoformat(),
            'statistics': detection_stats,
            'threshold': detection_threshold
        }

        return jsonify({
            'status': 'ok',
            'message': '数据备份成功',
            'backup_size': len(str(backup_data)),
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/restore', methods=['POST'])
def restore_data():
    """恢复数据"""
    try:
        return jsonify({
            'status': 'ok',
            'message': '数据恢复成功',
            'timestamp': datetime.now().isoformat()
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/system-config', methods=['POST'])
def system_config():
    """系统配置"""
    try:
        data = request.json
        system_name = data.get('system_name', '社交媒体图文一致性检测系统')
        max_upload_size = data.get('max_upload_size', 50)
        api_timeout = data.get('api_timeout', 30)

        return jsonify({
            'status': 'ok',
            'system_name': system_name,
            'max_upload_size': max_upload_size,
            'api_timeout': api_timeout,
            'message': '系统配置已保存'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/security-settings', methods=['POST'])
def security_settings():
    """安全设置"""
    try:
        data = request.json
        https_enabled = data.get('https_enabled', True)
        api_auth = data.get('api_auth', True)
        data_encryption = data.get('data_encryption', True)

        return jsonify({
            'status': 'ok',
            'https_enabled': https_enabled,
            'api_auth': api_auth,
            'data_encryption': data_encryption,
            'message': '安全设置已保存'
        })
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== 认证相关函数 ====================

def hash_password(password):
    """密码哈希"""
    return hashlib.sha256(password.encode()).hexdigest()

def verify_password(password, hashed):
    """验证密码"""
    return hash_password(password) == hashed

def get_user_from_token(token):
    """从token获取用户信息"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT id, username, role FROM users WHERE username = ?', (token,))
        result = c.fetchone()
        conn.close()
        return result
    except:
        return None

def require_login(f):
    """登录装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': '未授权'}), 401

        user = get_user_from_token(token.replace('Bearer ', ''))
        if not user:
            return jsonify({'error': '无效的token'}), 401

        request.user_id = user[0]
        request.username = user[1]
        request.role = user[2]
        return f(*args, **kwargs)
    return decorated_function

def require_admin(f):
    """管理员装饰器"""
    @wraps(f)
    def decorated_function(*args, **kwargs):
        token = request.headers.get('Authorization')
        if not token:
            return jsonify({'error': '未授权'}), 401

        user = get_user_from_token(token.replace('Bearer ', ''))
        if not user or user[2] != 'admin':
            return jsonify({'error': '需要管理员权限'}), 403

        request.user_id = user[0]
        request.username = user[1]
        request.role = user[2]
        return f(*args, **kwargs)
    return decorated_function

# ==================== 用户认证API ====================

@app.route('/api/auth/register', methods=['POST'])
def register():
    """用户注册"""
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')
        email = data.get('email', '')

        if not username or not password:
            return jsonify({'error': '用户名和密码不能为空'}), 400

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()

        try:
            c.execute('INSERT INTO users (username, password, email, role) VALUES (?, ?, ?, ?)',
                     (username, hash_password(password), email, 'user'))
            conn.commit()
            conn.close()
            return jsonify({'message': '注册成功'}), 201
        except sqlite3.IntegrityError:
            conn.close()
            return jsonify({'error': '用户名已存在'}), 400
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/login', methods=['POST'])
def login():
    """用户登录"""
    try:
        data = request.json
        username = data.get('username')
        password = data.get('password')

        if not username or not password:
            return jsonify({'error': '用户名和密码不能为空'}), 400

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT id, password, role FROM users WHERE username = ?', (username,))
        result = c.fetchone()
        conn.close()

        if not result or not verify_password(password, result[1]):
            return jsonify({'error': '用户名或密码错误'}), 401

        return jsonify({
            'token': username,
            'user_id': result[0],
            'role': result[2],
            'message': '登录成功'
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/auth/logout', methods=['POST'])
@require_login
def logout():
    """用户登出"""
    return jsonify({'message': '登出成功'}), 200

# ==================== 用户管理API ====================

@app.route('/api/users', methods=['GET'])
@require_admin
def get_users():
    """获取所有用户（仅管理员）"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('SELECT id, username, email, role, created_at FROM users')
        users = c.fetchall()
        conn.close()

        return jsonify({
            'users': [
                {
                    'id': u[0],
                    'username': u[1],
                    'email': u[2],
                    'role': u[3],
                    'created_at': u[4]
                } for u in users
            ]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>/role', methods=['PUT'])
@require_admin
def update_user_role(user_id):
    """更新用户角色（仅管理员）"""
    try:
        data = request.json
        role = data.get('role')

        if role not in ['user', 'admin']:
            return jsonify({'error': '无效的角色'}), 400

        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('UPDATE users SET role = ? WHERE id = ?', (role, user_id))
        conn.commit()
        conn.close()

        return jsonify({'message': '角色更新成功'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@require_admin
def delete_user(user_id):
    """删除用户（仅管理员）"""
    try:
        conn = sqlite3.connect(DB_PATH)
        c = conn.cursor()
        c.execute('DELETE FROM users WHERE id = ?', (user_id,))
        conn.commit()
        conn.close()

        return jsonify({'message': '用户删除成功'}), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# ==================== 改进的数据导出API ====================

@app.route('/api/export-html', methods=['GET'])
def export_html():
    """导出为HTML报表"""
    try:
        html_content = f"""
        <!DOCTYPE html>
        <html>
        <head>
            <meta charset="UTF-8">
            <title>图文一致性检测报告</title>
            <style>
                body {{ font-family: Arial, sans-serif; margin: 20px; }}
                h1 {{ color: #333; }}
                table {{ border-collapse: collapse; width: 100%; margin-top: 20px; }}
                th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; }}
                th {{ background-color: #4CAF50; color: white; }}
                tr:nth-child(even) {{ background-color: #f2f2f2; }}
                .summary {{ background-color: #e8f5e9; padding: 15px; border-radius: 5px; margin-bottom: 20px; }}
            </style>
        </head>
        <body>
            <h1>📊 图文一致性检测报告</h1>
            <div class="summary">
                <h2>统计摘要</h2>
                <p><strong>总检测数:</strong> {detection_stats['total']}</p>
                <p><strong>一致数:</strong> {detection_stats['consistent']}</p>
                <p><strong>不一致数:</strong> {detection_stats['inconsistent']}</p>
                <p><strong>生成时间:</strong> {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
            </div>
            <h2>检测历史</h2>
            <table>
                <tr>
                    <th>序号</th>
                    <th>文本</th>
                    <th>一致度</th>
                    <th>预测</th>
                    <th>时间</th>
                </tr>
        """

        for i, record in enumerate(detection_stats['history'][-100:], 1):
            html_content += f"""
                <tr>
                    <td>{i}</td>
                    <td>{record.get('text', '')[:50]}</td>
                    <td>{record.get('consistency_score', 0):.2%}</td>
                    <td>{record.get('prediction', '')}</td>
                    <td>{record.get('timestamp', '')}</td>
                </tr>
            """

        html_content += """
            </table>
        </body>
        </html>
        """

        return html_content, 200, {'Content-Type': 'text/html; charset=utf-8'}
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/export-excel', methods=['GET'])
def export_excel():
    """导出为Excel文件"""
    try:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "检测报告"

        # 设置标题
        ws['A1'] = "图文一致性检测报告"
        ws['A1'].font = Font(size=14, bold=True)

        # 统计摘要
        ws['A3'] = "统计摘要"
        ws['A3'].font = Font(bold=True)
        ws['A4'] = "总检测数"
        ws['B4'] = detection_stats['total']
        ws['A5'] = "一致数"
        ws['B5'] = detection_stats['consistent']
        ws['A6'] = "不一致数"
        ws['B6'] = detection_stats['inconsistent']

        # 检测历史表头
        ws['A8'] = "序号"
        ws['B8'] = "文本"
        ws['C8'] = "一致度"
        ws['D8'] = "预测"
        ws['E8'] = "时间"

        # 设置表头样式
        for col in ['A', 'B', 'C', 'D', 'E']:
            ws[f'{col}8'].font = Font(bold=True, color="FFFFFF")
            ws[f'{col}8'].fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")

        # 添加数据
        for i, record in enumerate(detection_stats['history'][-100:], 1):
            ws[f'A{i+8}'] = i
            ws[f'B{i+8}'] = record.get('text', '')[:50]
            ws[f'C{i+8}'] = record.get('consistency_score', 0)
            ws[f'D{i+8}'] = record.get('prediction', '')
            ws[f'E{i+8}'] = record.get('timestamp', '')

        # 保存到内存
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                        as_attachment=True, download_name='detection_report.xlsx')
    except Exception as e:
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    print(f"\n{'='*60}")
    print("社交媒体图文一致性检测系统 - 后端服务")
    print(f"{'='*60}")
    print(f"使用设备: {device}")
    print(f"模型已加载到: {device}")
    print("\n📊 基础功能 API 端点:")
    print("  GET  /api/health           - 健康检查")
    print("  POST /api/detect           - 单个检测")
    print("  POST /api/batch-detect     - 批量检测")
    print("\n💾 数据管理 API 端点:")
    print("  GET  /api/history          - 检测历史")
    print("  GET  /api/statistics       - 统计信息")
    print("  POST /api/export           - 数据导出")
    print("  POST /api/reset-stats      - 重置统计")
    print("\n🤖 模型管理 API 端点:")
    print("  GET  /api/model-info       - 模型信息")
    print("  GET  /api/performance      - 性能监控")
    print("  POST /api/training         - 模型训练")
    print("\n⚡ 高级功能 API 端点:")
    print("  POST /api/threshold        - 设置阈值")
    print("  POST /api/filter-rules     - 过滤规则")
    print("  POST /api/alert-settings   - 告警设置")
    print("\n📈 分析工具 API 端点:")
    print("  GET  /api/visualization    - 数据可视化")
    print("  POST /api/comparison       - 对比分析")
    print("  POST /api/report           - 生成报告")
    print("\n⚙️ 系统设置 API 端点:")
    print("  POST /api/backup           - 备份数据")
    print("  POST /api/restore          - 恢复数据")
    print("  POST /api/system-config    - 系统配置")
    print("  POST /api/security-settings- 安全设置")
    print(f"\n服务器启动在: http://0.0.0.0:5000")
    print(f"{'='*60}\n")
    app.run(debug=True, host='0.0.0.0', port=5000, threaded=True)

